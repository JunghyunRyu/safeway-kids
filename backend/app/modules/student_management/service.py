from __future__ import annotations

import io
import uuid
from datetime import UTC, date, datetime
from typing import TYPE_CHECKING

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

if TYPE_CHECKING:
    from app.modules.auth.models import User

from app.common.exceptions import ConflictError, NotFoundError
from app.middleware.consent import require_consent
from app.modules.student_management.models import Enrollment, Student
from app.modules.student_management.schemas import (
    BulkUploadRowResult,
    EnrollmentCreateRequest,
    StudentCreateRequest,
    StudentUpdateRequest,
)


async def create_student(
    db: AsyncSession, guardian_id: uuid.UUID, request: StudentCreateRequest
) -> Student:
    # Consent check is done at registration time (consent must exist before creating student data)
    student = Student(
        guardian_id=guardian_id,
        name=request.name,
        date_of_birth=request.date_of_birth,
        grade=request.grade,
    )
    db.add(student)
    await db.flush()
    return student


async def get_student(db: AsyncSession, student_id: uuid.UUID) -> Student:
    stmt = select(Student).where(Student.id == student_id, Student.deleted_at.is_(None))
    result = await db.execute(stmt)
    student = result.scalar_one_or_none()
    if not student:
        raise NotFoundError(detail="학생을 찾을 수 없습니다")
    return student


async def list_students_by_academy(
    db: AsyncSession, admin_id: uuid.UUID, skip: int = 0, limit: int = 20
) -> dict:
    """학원 관리자 소속 학원의 전체 학생 목록 (페이지네이션)."""
    from sqlalchemy import func as sa_func

    from app.modules.academy_management.models import Academy
    academy_stmt = select(Academy).where(Academy.admin_id == admin_id)
    academy_result = await db.execute(academy_stmt)
    academy = academy_result.scalar_one_or_none()
    if not academy:
        return {"items": [], "total": 0}

    base = (
        select(Student)
        .join(Enrollment, Enrollment.student_id == Student.id)
        .where(
            Enrollment.academy_id == academy.id,
            Enrollment.withdrawn_at.is_(None),
            Student.deleted_at.is_(None),
        )
    )
    count_stmt = select(sa_func.count()).select_from(base.subquery())
    total = (await db.execute(count_stmt)).scalar() or 0

    items_stmt = base.order_by(Student.created_at.desc()).offset(skip).limit(limit)
    result = await db.execute(items_stmt)
    return {"items": list(result.scalars().all()), "total": total}


async def list_all_students(db: AsyncSession, skip: int = 0, limit: int = 20) -> dict:
    """플랫폼 관리자: 전체 학생 목록 (페이지네이션)."""
    from sqlalchemy import func as sa_func

    base = select(Student).where(Student.deleted_at.is_(None))
    count_stmt = select(sa_func.count()).select_from(base.subquery())
    total = (await db.execute(count_stmt)).scalar() or 0

    items_stmt = base.order_by(Student.created_at.desc()).offset(skip).limit(limit)
    result = await db.execute(items_stmt)
    return {"items": list(result.scalars().all()), "total": total}


async def list_students_by_guardian(
    db: AsyncSession, guardian_id: uuid.UUID, skip: int = 0, limit: int = 20
) -> dict:
    """학부모: 자녀 목록 (페이지네이션)."""
    from sqlalchemy import func as sa_func

    base = select(Student).where(
        Student.guardian_id == guardian_id, Student.deleted_at.is_(None)
    )
    count_stmt = select(sa_func.count()).select_from(base.subquery())
    total = (await db.execute(count_stmt)).scalar() or 0

    items_stmt = base.order_by(Student.created_at.desc()).offset(skip).limit(limit)
    result = await db.execute(items_stmt)
    return {"items": list(result.scalars().all()), "total": total}


async def update_student(
    db: AsyncSession,
    student_id: uuid.UUID,
    current_user: "User",
    request: StudentUpdateRequest,
) -> Student:
    from app.modules.auth.models import UserRole

    student = await get_student(db, student_id)

    # Platform admin and academy admin can edit any student; parent only their own
    if current_user.role == UserRole.PARENT and student.guardian_id != current_user.id:
        from app.common.exceptions import ForbiddenError
        raise ForbiddenError(detail="본인의 자녀만 수정할 수 있습니다")

    if request.name is not None:
        student.name = request.name
    if request.grade is not None:
        student.grade = request.grade

    await db.flush()
    return student


async def deactivate_student(
    db: AsyncSession, student_id: uuid.UUID, current_user: "User"
) -> Student:
    from app.modules.auth.models import UserRole

    student = await get_student(db, student_id)

    # Platform admin and academy admin can deactivate any student; parent only their own
    if current_user.role == UserRole.PARENT and student.guardian_id != current_user.id:
        from app.common.exceptions import ForbiddenError
        raise ForbiddenError(detail="본인의 자녀만 비활성화할 수 있습니다")

    student.is_active = False
    student.deleted_at = datetime.now(UTC)
    await db.flush()
    return student


async def enroll_student(
    db: AsyncSession,
    student_id: uuid.UUID,
    guardian_id: uuid.UUID,
    request: EnrollmentCreateRequest,
) -> Enrollment:
    student = await get_student(db, student_id)
    if student.guardian_id != guardian_id:
        from app.common.exceptions import ForbiddenError
        raise ForbiddenError(detail="본인의 자녀만 등록할 수 있습니다")

    # Check consent
    await require_consent(db, guardian_id, student_id)

    # Check duplicate
    stmt = select(Enrollment).where(
        Enrollment.student_id == student_id,
        Enrollment.academy_id == request.academy_id,
        Enrollment.withdrawn_at.is_(None),
    )
    result = await db.execute(stmt)
    if result.scalar_one_or_none():
        raise ConflictError(detail="이미 등록된 학원입니다")

    enrollment = Enrollment(
        student_id=student_id,
        academy_id=request.academy_id,
    )
    db.add(enrollment)
    await db.flush()
    return enrollment


async def list_enrollments(
    db: AsyncSession, student_id: uuid.UUID, guardian_id: uuid.UUID
) -> list[Enrollment]:
    """List active enrollments for a student (parent only)."""
    student = await get_student(db, student_id)
    if student.guardian_id != guardian_id:
        from app.common.exceptions import ForbiddenError
        raise ForbiddenError(detail="본인의 자녀 등록만 조회할 수 있습니다")

    stmt = (
        select(Enrollment)
        .where(Enrollment.student_id == student_id, Enrollment.withdrawn_at.is_(None))
        .order_by(Enrollment.enrolled_at.desc())
    )
    result = await db.execute(stmt)
    return list(result.scalars().all())


async def withdraw_enrollment(
    db: AsyncSession,
    enrollment_id: uuid.UUID,
    guardian_id: uuid.UUID,
) -> Enrollment:
    stmt = select(Enrollment).where(Enrollment.id == enrollment_id)
    result = await db.execute(stmt)
    enrollment = result.scalar_one_or_none()
    if not enrollment:
        raise NotFoundError(detail="등록 정보를 찾을 수 없습니다")

    # Verify ownership via student
    student = await get_student(db, enrollment.student_id)
    if student.guardian_id != guardian_id:
        from app.common.exceptions import ForbiddenError
        raise ForbiddenError(detail="본인의 자녀 등록만 철회할 수 있습니다")

    enrollment.withdrawn_at = datetime.now(UTC)
    await db.flush()
    return enrollment


# Column name mapping: Korean header → field name
_COLUMN_MAP: dict[str, str] = {
    "이름": "name",
    "name": "name",
    "생년월일": "date_of_birth",
    "date_of_birth": "date_of_birth",
    "학년": "grade",
    "grade": "grade",
    "보호자전화번호": "guardian_phone",
    "guardian_phone": "guardian_phone",
    "학원명": "academy_name",
    "academy_name": "academy_name",
    "academy_id": "academy_id",
}

MAX_BULK_ROWS = 500


def _parse_date(value: str | datetime | date | None) -> date | None:
    """Parse a date value from Excel (could be datetime object or string)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


async def bulk_upload_students(
    db: AsyncSession,
    admin_id: uuid.UUID,
    file_bytes: bytes,
) -> list[BulkUploadRowResult]:
    """Parse an Excel file and create students with enrollments in bulk.

    The admin must own an academy. Each row creates a Student (with guardian
    looked up by phone) and an Enrollment to the admin's academy.
    Partial success is allowed — successful rows are committed even if others fail.
    """
    from openpyxl import load_workbook

    from app.modules.academy_management.models import Academy

    # Resolve the admin's academy
    academy_stmt = select(Academy).where(Academy.admin_id == admin_id)
    academy_result = await db.execute(academy_stmt)
    academy = academy_result.scalar_one_or_none()
    if not academy:
        return [BulkUploadRowResult(row=0, status="error", message="관리자에게 연결된 학원이 없습니다")]

    # Load workbook
    try:
        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return [BulkUploadRowResult(row=0, status="error", message="엑셀 파일을 읽을 수 없습니다")]

    ws = wb.active
    if ws is None:
        return [BulkUploadRowResult(row=0, status="error", message="시트를 찾을 수 없습니다")]

    rows_iter = ws.iter_rows(values_only=True)

    # Parse header row
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        return [BulkUploadRowResult(row=0, status="error", message="빈 파일입니다")]

    headers: list[str | None] = []
    for h in raw_headers:
        key = str(h).strip().lower() if h else None
        if key and key in _COLUMN_MAP:
            headers.append(_COLUMN_MAP[key])
        else:
            headers.append(None)

    # Ensure required columns exist
    required = {"name", "date_of_birth", "guardian_phone"}
    present = {h for h in headers if h is not None}
    missing = required - present
    if missing:
        return [
            BulkUploadRowResult(
                row=0,
                status="error",
                message=f"필수 컬럼 누락: {', '.join(sorted(missing))}",
            )
        ]

    results: list[BulkUploadRowResult] = []

    for row_idx, row_values in enumerate(rows_iter, start=2):
        if row_idx - 1 > MAX_BULK_ROWS:
            results.append(
                BulkUploadRowResult(row=row_idx, status="error", message="최대 행 수(500)를 초과했습니다")
            )
            break

        # Build a dict from the row
        row_data: dict[str, str | None] = {}
        for col_idx, val in enumerate(row_values):
            if col_idx < len(headers) and headers[col_idx]:
                row_data[headers[col_idx]] = str(val).strip() if val is not None else None

        # Skip entirely empty rows
        if all(v is None or v == "" for v in row_data.values()):
            continue

        # Validate required fields
        name = row_data.get("name")
        dob_raw = row_values[headers.index("date_of_birth")] if "date_of_birth" in headers else None
        guardian_phone = row_data.get("guardian_phone")
        grade = row_data.get("grade")

        if not name:
            results.append(BulkUploadRowResult(row=row_idx, status="error", message="이름이 비어있습니다"))
            continue

        dob = _parse_date(dob_raw)
        if not dob:
            results.append(
                BulkUploadRowResult(row=row_idx, status="error", message="생년월일 형식이 올바르지 않습니다")
            )
            continue

        if not guardian_phone:
            results.append(
                BulkUploadRowResult(row=row_idx, status="error", message="보호자전화번호가 비어있습니다")
            )
            continue

        # Normalise phone: remove dashes/spaces
        guardian_phone = guardian_phone.replace("-", "").replace(" ", "")

        # Look up guardian by phone
        from app.modules.auth.models import User, UserRole

        guardian_stmt = select(User).where(
            User.phone == guardian_phone,
            User.role == UserRole.PARENT,
            User.is_active.is_(True),
        )
        guardian_result = await db.execute(guardian_stmt)
        guardian = guardian_result.scalar_one_or_none()
        if not guardian:
            results.append(
                BulkUploadRowResult(
                    row=row_idx,
                    status="error",
                    message=f"보호자 전화번호({guardian_phone})에 해당하는 학부모를 찾을 수 없습니다",
                )
            )
            continue

        # Duplicate check: same name + dob + guardian
        dup_stmt = select(Student).where(
            Student.name == name,
            Student.date_of_birth == dob,
            Student.guardian_id == guardian.id,
            Student.deleted_at.is_(None),
        )
        dup_result = await db.execute(dup_stmt)
        existing_student = dup_result.scalar_one_or_none()

        if existing_student:
            # Student exists — check if already enrolled in this academy
            enroll_dup_stmt = select(Enrollment).where(
                Enrollment.student_id == existing_student.id,
                Enrollment.academy_id == academy.id,
                Enrollment.withdrawn_at.is_(None),
            )
            enroll_dup_result = await db.execute(enroll_dup_stmt)
            if enroll_dup_result.scalar_one_or_none():
                results.append(
                    BulkUploadRowResult(
                        row=row_idx, status="error", message="이미 등록된 학생입니다 (중복)"
                    )
                )
                continue
            # Enroll existing student
            enrollment = Enrollment(student_id=existing_student.id, academy_id=academy.id)
            db.add(enrollment)
            await db.flush()
            results.append(
                BulkUploadRowResult(
                    row=row_idx, status="success", message=f"기존 학생 '{name}' 학원 등록 완료"
                )
            )
            continue

        # Create new student
        student = Student(
            guardian_id=guardian.id,
            name=name,
            date_of_birth=dob,
            grade=grade,
        )
        db.add(student)
        await db.flush()

        # Create enrollment
        enrollment = Enrollment(student_id=student.id, academy_id=academy.id)
        db.add(enrollment)
        await db.flush()

        results.append(
            BulkUploadRowResult(row=row_idx, status="success", message=f"학생 '{name}' 등록 완료")
        )

    return results
